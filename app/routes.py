from flask import Blueprint, render_template, request, redirect, url_for, jsonify
from . import db
from .models import Student
from sqlalchemy import or_

bp = Blueprint("main", __name__)

@bp.route("/")
def index():
    authorizations = Student.query.all()
    return render_template("index.html", authorizations=authorizations)

@bp.route("/searchStudents/")
def searchStudents():
    searchInput = request.args.get("query", "").strip()

    if not searchInput:
        return jsonify({"error": "Debe ingresar un valor de búsqueda"}), 400

    # Buscar por nombre, número de cédula, teléfono o email
    existStudent = Student.query.filter(
        or_(
            Student.name.ilike(f"%{searchInput}%"),
            Student.id_number.ilike(f"%{searchInput}%"),
            Student.phone_number.ilike(f"%{searchInput}%"),
            Student.email.ilike(f"%{searchInput}%")
        )
    ).all()

    if not existStudent:
        return jsonify({"message": "No se encontraron estudiantes"}), 404

    # Convertir los resultados en un formato JSON amigable
    results = [
        {
            "id": s.id,
            "name": s.name,
            "id_number": s.id_number,
            "area":s.area,
            "phone_number": s.phone_number,
            "email": s.email
        }
        for s in existStudent
    ]

    return jsonify(results), 200

@bp.route("/addAuthorization/", methods=["POST"])
def confirmUseDiscount():
    return 'Data'

from openpyxl import load_workbook

@bp.route("/migrate/")
def migrate():
    errores = []
    guardados = 0
    ruta_archivo = "datos.xlsx"

    try:
        wb = load_workbook(ruta_archivo)
        hoja = wb.active
    except Exception as e:
        return jsonify({"error": f"No se pudo abrir el archivo Excel: {e}"}), 500

    for fila in hoja.iter_rows(min_row=2, values_only=True):  # saltar encabezado
        if not fila or all(celda is None for celda in fila):
            continue

        name, idNumber, phoneNumber, area, email = fila[:5]

        if not idNumber:
            errores.append("Fila sin número de cédula, omitida.")
            continue

        try:
            new_student = Student(
                name=name or "",
                id_number=idNumber,
                phone_number=phoneNumber or "",
                email=email or "",
                area=area or ""
            )
            db.session.add(new_student)
            db.session.commit()
            guardados += 1
        except Exception as e:
            db.session.rollback()  # importante para limpiar la sesión rota
            errores.append(f"Error con {idNumber} ({email}): {e}")

    return jsonify({
        "mensaje": f"Migración completada: {guardados} estudiantes guardados.",
        "errores": errores
    }), 200
